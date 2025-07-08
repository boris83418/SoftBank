import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import create_engine
import sys
import logging
import os  # Add os module to handle paths
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import calendar

# Configure logging
log_filename = f"logfile_{datetime.now().strftime('%Y-%m-%d')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)


# Connect to MSSQL
def connect_to_db(server, database):
    try:
        conn_str = f"mssql+pyodbc://@{server}/{database}?driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes"
        engine = create_engine(conn_str)
        return engine
    except Exception as e:
        logging.error(f"資料庫連接失敗: {e}")
        raise


# Fetch data
def fetch_data():
    try:
        engine = connect_to_db('jpdejitdev01', 'ITQAS2')
        queries = {
            'factory_data': "SELECT Part_No, eta_FLTC, Qty, Status FROM SoftBank_Data_FactoryShipment",
            'order_data': "SELECT Product_Name, COALESCE(Actual_shipment_Date, Estimated_Shipment_Date) AS Shipment_Date, Quantity, Quotation_status FROM SoftBank_Data_Orderinfo",
            'product_data': "SELECT Delta_PartNO AS Part_No, [Month-End_SAP_Inventory], Model FROM SoftBank_Data_Productinfo"
        }
        return {name: pd.read_sql(query, engine) for name, query in queries.items()}
    except Exception as e:
        logging.error(f"Error fetching data: {e}")
        raise

def calculate_inventory(factory_data, order_data, product_data):
    try:
        # print("=== 開始庫存計算調試 ===")
        
        # Clean data
        factory_data['eta_FLTC'] = pd.to_datetime(factory_data['eta_FLTC']).dt.date
        order_data['Shipment_Date'] = pd.to_datetime(order_data['Shipment_Date']).dt.date

        # Perform inventory calculation(half year count)
        start_date = datetime.today().replace(day=1).date()
        end_date = start_date + timedelta(days=180)
        date_range = pd.date_range(start=start_date, end=end_date).date

        # print(f"日期範圍: {start_date} 到 {end_date}")

        # 過濾訂單數據
        order_data = order_data[
            (order_data['Shipment_Date'] >= start_date) & 
            (~order_data['Quotation_status'].isin(['quotation', 'cancel','confirming','double cancel']))
        ]

        # 獲取所有產品列表
        products = product_data['Part_No'].unique()
        # print(f"總產品數量: {len(products)}")
        
        # 創建庫存DataFrame
        inventory = pd.DataFrame(index=date_range, columns=products, data=0.0)
        # print(f"初始庫存DataFrame形狀: {inventory.shape}")

        # 設定初始庫存
        # print("\n=== 設定初始庫存 ===")
        
        # 清理product_data，移除重複和空值
        clean_product_data = product_data.dropna(subset=['Part_No', 'Month-End_SAP_Inventory'])
        clean_product_data = clean_product_data.drop_duplicates(subset=['Part_No'], keep='first')
        
        # print(f"清理後的product_data數量: {len(clean_product_data)}")
        
        # 逐個設定初始庫存
        for product in products:
            matching_rows = clean_product_data[clean_product_data['Part_No'] == product]
            
            if len(matching_rows) > 0:
                initial_qty = matching_rows['Month-End_SAP_Inventory'].iloc[0]
                # 確保是數值類型
                try:
                    initial_qty = float(initial_qty) if pd.notna(initial_qty) else 0.0
                    inventory.loc[start_date, product] = initial_qty
                except (ValueError, TypeError):
                    inventory.loc[start_date, product] = 0.0
            else:
                inventory.loc[start_date, product] = 0.0
        
        # 驗證第一天庫存設定（保留關鍵驗證）
        first_day_stock = inventory.loc[start_date]
        non_zero_count = (first_day_stock > 0).sum()
        print(f"第一天有庫存的產品數量: {non_zero_count}")
        
        # 處理工廠數據
        # print("\n=== 處理工廠進貨數據 ===")
        valid_factory_data = factory_data[factory_data['Part_No'].isin(products)]
        # print(f"有效工廠數據筆數: {len(valid_factory_data)}")
        
        if not valid_factory_data.empty:
            factory_agg = valid_factory_data.groupby(['eta_FLTC', 'Part_No'])['Qty'].sum().unstack(fill_value=0)
            factory_agg = factory_agg.reindex(columns=products, index=date_range, fill_value=0)
        else:
            factory_agg = pd.DataFrame(0, index=date_range, columns=products)
        
        # 處理訂單數據
        # print("\n=== 處理訂單出貨數據 ===")
        # print(f"有效訂單數據筆數: {len(order_data)}")
        
        if not order_data.empty:
            order_agg = order_data.groupby(['Shipment_Date', 'Product_Name'])['Quantity'].sum().unstack(fill_value=0)
            order_agg = order_agg.reindex(columns=products, index=date_range, fill_value=0)
        else:
            order_agg = pd.DataFrame(0, index=date_range, columns=products)

        # 庫存計算邏輯
        # print("\n=== 執行庫存計算 ===")
        
        # 逐日計算庫存
        for i, current_date in enumerate(date_range):
            if i == 0:
                # 第一天：初始庫存 + 當天進貨 - 當天出貨
                daily_in = factory_agg.loc[current_date]
                daily_out = order_agg.loc[current_date]
                inventory.loc[current_date] = inventory.loc[current_date] + daily_in - daily_out
            else:
                # 其他天：前一天庫存 + 當天進貨 - 當天出貨
                prev_date = date_range[i-1]
                daily_in = factory_agg.loc[current_date]
                daily_out = order_agg.loc[current_date]
                inventory.loc[current_date] = inventory.loc[prev_date] + daily_in - daily_out
        
        # print("=== 庫存計算完成 ===\n")
        
        return inventory
        
    except Exception as e:
        logging.error(f"Error during inventory calculation: {e}")
        raise

# Export results to Excel
def export_to_excel(inventory, product_data):
    try:
        print("\n開始處理庫存數據...")

        # ===== 讀取對應表 Excel =====
        mapping_file = r"\\jpdejstcfs01\STC_share\JP IT\STC SBK 仕分けリスト\IT system\part_mapping.xlsx"
        one_to_one_df = pd.read_excel(mapping_file, sheet_name='OneToOne')
        many_to_one_df = pd.read_excel(mapping_file, sheet_name='ManyToOne')
        excluded_df = pd.read_excel(mapping_file, sheet_name='Exclude')

        # ===== 排除不計算料號 =====
        excluded_parts = excluded_df['Excluded_Part_No'].dropna().tolist()
        for part_no in excluded_parts:
            if part_no in inventory.columns:
                inventory.drop(columns=[part_no], inplace=True)
                print(f"  ✓ 已排除: {part_no}")

        # ===== 一對一合併 =====
        for _, row in one_to_one_df.iterrows():
            free_part, main_part = row['Free_Part_No'], row['Main_Part_No']
            if free_part in inventory.columns:
                if main_part in inventory.columns:
                    inventory[main_part] += inventory[free_part]
                else:
                    inventory.rename(columns={free_part: main_part}, inplace=True)
                inventory.drop(columns=[free_part], inplace=True, errors='ignore')

        # ===== 多對一合併 =====
        for main_part, group_df in many_to_one_df.groupby('Main_Part_No'):
            for alias_part in group_df['Alias_Part_No']:
                if alias_part in inventory.columns:
                    if main_part in inventory.columns:
                        inventory[main_part] += inventory[alias_part]
                    else:
                        inventory[main_part] = inventory[alias_part]
                    inventory.drop(columns=[alias_part], inplace=True, errors='ignore')

        # ===== 插入 Model =====
        inventory_transposed = inventory.T
        part_no_model_mapping = product_data.set_index('Part_No')['Model'].to_dict()
        inventory_transposed.insert(0, 'Model', inventory_transposed.index.map(part_no_model_mapping))

        # ===== 檔名與儲存路徑 =====
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        file_name = f"Daily_Inventory_Simulate_{timestamp}.xlsx"
        save_path = r"\\jpdejstcfs01\STC_share\JP IT\STC SBK 仕分けリスト\IT system\Report"
        full_path = os.path.join(save_path, file_name)

        inventory_transposed.to_excel(full_path, index=True)
        wb = load_workbook(full_path)
        ws = wb.active

        # ===== 格式設定 =====
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col_cells in ws.iter_cols(min_row=1, max_row=1):
            for cell in col_cells:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

        def get_month_end_columns():
            month_end_cols = []
            for col in range(3, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and isinstance(cell_value, datetime):
                    last_day = calendar.monthrange(cell_value.year, cell_value.month)[1]
                    if cell_value.day == last_day:
                        month_end_cols.append(col)
            return month_end_cols

        month_end_cols = get_month_end_columns()

        light_gray_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        negative_font = Font(color="FF0000")
        negative_bold_font = Font(color="FF0000", bold=True)
        bold_font = Font(bold=True)

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = light_gray_fill

                if isinstance(cell.value, (int, float)):
                    cell.alignment = right_align
                    cell.number_format = "#,##0"
                    is_negative = cell.value < 0
                    is_month_end = col in month_end_cols
                    if is_negative and is_month_end:
                        cell.font = negative_bold_font
                    elif is_negative:
                        cell.font = negative_font
                    elif is_month_end:
                        cell.font = bold_font
                else:
                    cell.alignment = left_align

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        ws.freeze_panes = "A2"
        wb.save(full_path)

        print(f"\n報表匯出成功: {full_path}")

    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")
        raise
# Main function
def main()-> int:
    try:
        # Fetch data
        data = fetch_data()
        factory_data = data['factory_data']
        order_data = data['order_data']
        product_data = data['product_data']

        # Calculate daily inventory
        inventory = calculate_inventory(factory_data, order_data, product_data)
        
        # Export to Excel
        export_to_excel(inventory,product_data)
        return 0
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        sys.exit(1)  # Exit the program with a non-zero status code
        return 1

#直接執行的Py的時候
if __name__ == "__main__":
    try:
        # Fetch data
        data = fetch_data()
        factory_data = data['factory_data']
        order_data = data['order_data']
        product_data = data['product_data']

        # Calculate daily inventory
        inventory = calculate_inventory(factory_data, order_data, product_data)
        
        # Export to Excel
        export_to_excel(inventory,product_data)
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        sys.exit(1)  # Exit the program with a non-zero status code